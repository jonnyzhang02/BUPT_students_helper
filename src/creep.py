def creeper():
    import re
    # import requests
    import time
    import selenium
    from selenium import webdriver
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.reader.excel import load_workbook
    import csv
    from selenium.webdriver.support.wait import WebDriverWait #导入显性等待的包
    from selenium.webdriver.support import expected_conditions as EC #判断所需要的元素是否已经被加载出来
    from selenium.webdriver.common.by import By #定位

    url = 'https://webvpn.bupt.edu.cn'  # 北邮人首页网址
    byr_page = webdriver.Edge()  # 选择浏览器
    byr_page.get(url)  # 获取页面
    byr_page.maximize_window()
    WebDriverWait(byr_page,60,0.5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div/div/div/div[3]/div/div[2]/div/div[1]/div/div[1]/div/div[1]/a')))
    botton1 = byr_page.find_element('xpath','/html/body/div/div/div/div/div[3]/div/div[2]/div/div[1]/div/div[1]/div/div[1]/a')
    botton1.click()

    byr_page.switch_to.window(byr_page.window_handles[1])
    WebDriverWait(byr_page,60,0.5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[3]/div[2]/div/div[1]/a/p')))
    botton2 = byr_page.find_element('xpath','/html/body/div[2]/div[3]/div[2]/div/div[1]/a/p')
    botton2.click()

    byr_page.switch_to.window(byr_page.window_handles[2])
    WebDriverWait(byr_page,60,0.5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[3]/div/div[2]/iframe')))

    fm1 = byr_page.find_element('xpath','/html/body/div[3]/div/div[2]/iframe')
    byr_page.switch_to.frame(fm1)

    button3 = byr_page.find_element('xpath','/html/body/div/div[2]/div[1]/div[2]/div/div/div[1]/p')
    button3.click()

    byr_page.switch_to.window(byr_page.window_handles[2])
    fm2 = byr_page.find_element('xpath','/html/body/div[3]/div/div[2]/iframe[2]')
    byr_page.switch_to.frame(fm2)


    time.sleep(5)

    kebiao = []
    for k in range(16):
        week_path = str('/html/body/div/div/form[2]/div[2]/select[1]/option['+str(k+2)+']')
        week = byr_page.find_element('xpath', week_path)
        week.click()
        time.sleep(2)
        class_data = []
        for i in range(7):
            day_data = []
            for j in range(14):
                classes_path = str('/html/body/div/div/form[2]/table[1]/tbody/tr['+str(j+2)+']/td['+str(i+1)+']/div[2]')
                try:
                    data = byr_page.find_element('xpath',classes_path)
                    day_data.append(data.text)
                except:
                    day_data.append(' ')
            class_data.append(day_data)
        kebiao.append(class_data)

    byr_page.close()

    kecheng = openpyxl.Workbook()
    kecheng.active.title_path = "课程表"
    kecheng.create_sheet('课程表')
    sheet1 = kecheng['课程表']

    for k in range(len(kebiao)):
        for i in range(7):
            for j in range(14):
                sheet1.cell(j+1, i+1+(k*7)).value = kebiao[k][i][j]


    kecheng.save("课程表.xlsx")
