import _thread
import datetime
import re
import time
from tkinter import *
import pygame
import openpyxl


def showMessage(tex, music):
    # show reminder message window
    root = Tk()  # 建立根窗口
    play_music(music)
    root.withdraw()  # hide window
    # 获取屏幕的宽度和高度，并且在高度上考虑到底部的任务栏，为了是弹出的窗口在屏幕中间
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight() - 100
    root.resizable(False, False)
    # 添加组件
    root.title("温馨提示")
    frame = Frame(root, relief=RIDGE, borderwidth=3)
    frame.pack(fill=BOTH, expand=1)  # pack() 放置组件若没有则组件不会显示
    # 窗口显示的文字、并设置字体、字号
    label = Label(frame, text=tex, font=('微软雅黑', 20, 'bold'))
    label.pack(fill=BOTH, expand=1)
    # 按钮的设置
    button = Button(frame, text="我知道了", font="Cooper -25 bold", fg="purple", command=shut_down)
    button.pack(side=BOTTOM)

    root.update_idletasks()
    root.deiconify()  # now the window size was calculated
    root.withdraw()  # hide the window again 防止窗口出现被拖动的感觉 具体原理未知？
    root.geometry('%sx%s+%s+%s' % (root.winfo_width() + 10, root.winfo_height() + 10,
                                   int((screenwidth - root.winfo_width()) / 2),
                                   int((screenheight - root.winfo_height()) / 2)))
    root.deiconify()
    root.mainloop()


def play_music(music):
    pygame.mixer.init()
    pygame.mixer.music.load(music)
    pygame.mixer.music.play()


def shut_down():
    pygame.mixer.music.stop()


def read_class_info():
    workbook = openpyxl.load_workbook('课程表.xlsx')
    sheet = workbook['课程表']
    class_info_ = [[]]
    for r in range(14):
        for c in range(112):
            temp = sheet.cell(row=r + 1, column=c + 1).value
            one = temp.find('\n')
            two = temp.find('\n', one + 1)
            if not one == -1:
                temp = temp[:two]
            class_info_[r].append(temp)
        class_info_.append([])
    return class_info_


def get_today():
    date0 = '2022-02-28'
    d0 = datetime.datetime.strptime(date0, '%Y-%m-%d')
    date1 = time.strftime("%Y-%m-%d", time.localtime())
    d1 = datetime.datetime.strptime(date1, '%Y-%m-%d')
    today = int((d1 - d0).days)
    return today


def get_class_alarm(ahead=0):
    class_info = read_class_info()
    have_class = class_info
    for line in range(14):
        for row in range(len(class_info[0])):
            if class_info[line][row] == ' ':
                have_class[line][row] = 0
            else:
                have_class[line][row] = 1

    class_times = ["08:00", "08:50", "09:50", "10:40", "11:30", "13:00", "13:50", "14:45", "15:40", "16:35", "17:25",
                   "18:30", "19:20", "20:10"]
    alarm_times = []
    today = get_today()
    for i in range(14):
        if have_class[i][today]:
            alarm_times.append(class_times[i])

    ahead = 60 * ahead
    for i in range(len(alarm_times)):
        date = time.strftime("%Y-%m-%d", time.localtime())
        alarm_times[i] = time.mktime(time.strptime(date + ' ' + alarm_times[i], "%Y-%m-%d %H:%M")) - ahead

    return alarm_times


def get_to_do_alarm(ahead=0):
    workbook = openpyxl.load_workbook('待办事项.xlsx')
    sheet = workbook['待办事项']
    to_do_info_ = [[]]
    for r in range(20):
        for c in range(112):
            temp = sheet.cell(row=r + 1, column=c + 1).value
            to_do_info_[r].append(temp)
        to_do_info_.append([])

    alarm_times = []
    alarm_things = []
    today = get_today()
    for i in range(20):
        if not to_do_info_[i][today] == ' ':
            to_do_info_[i][today]=str(to_do_info_[i][today]).lstrip()
            print(to_do_info_[i][today])
            p = re.compile(r'\d\d:\d\d')
            temp = p.findall(to_do_info_[i][today])[0]
            date = time.strftime("%Y-%m-%d", time.localtime())
            alarm_times.append(time.mktime(time.strptime(date + ' ' + temp, "%Y-%m-%d %H:%M")))
            alarm_things.append(to_do_info_[i][today][6:])

    return [alarm_times, alarm_things]


def mainloop(ahead, music):
    class_alarms = get_class_alarm(ahead)
    to_do_info = get_to_do_alarm()
    to_do_alarms = to_do_info[0]
    to_do_things = to_do_info[1]
    print(class_alarms)
    print(to_do_alarms)
    while 1:
        time.sleep(1)
        for each in class_alarms:
            if each == int(time.time()):
                showMessage("现在距离上课还有{}分钟，请注意！！！！！".format(ahead), music)
        for i in range(len(to_do_alarms)):
            if to_do_alarms[i] == int(time.time()):
                showMessage("快到\n{}\n的时间了，请注意！！！！！".format(to_do_things[i]), music)


def alarm_system(ahead, music):
    _thread.start_new_thread(mainloop, (ahead, music))

