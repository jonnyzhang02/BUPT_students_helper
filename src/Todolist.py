from tkinter import *
import openpyxl
from PIL import Image, ImageTk
import tkinter
import tkinter.messagebox as msgbox
import os
import copy
from tkinter import BooleanVar
import datetime
import time

def pop_up_box(w_num, d_num):
    print("周数：{}".format(w_num))
    # 调用至read_or_write做写入工作
    # 使用tkinter弹出输入框输入数字, 具有确定输入和清除功能
    def input_int():
        nonlocal event, ddl
        if var1.get() and var2.get():
            event = var1.get()
            ddl = var2.get()
            # 保存至文件
            workbook = openpyxl.load_workbook('待办事项.xlsx')
            ws = workbook['待办事项']
            r = 1
            for i in range(20):
                if ws.cell(row=i + 1, column=(w_num - 1) * 7 + d_num).value == ' ':
                    r = i + 1
                    break

            ws.cell(row=r, column=(w_num - 1) * 7 + d_num).value += ddl + '\n' + event
            workbook.save('待办事项.xlsx')  # 保存变更

            # exit()

        else:
            msgbox.showinfo('温馨提⽰', '请不要输入空白信息')

    def input_clear():
        nonlocal event, ddl
        var1.set('')
        var2.set('')
        event, ddl = '', ''

    event, ddl = '', ''
    z = tkinter.Toplevel()  # 弹出框框名
    z.geometry('270x150+500+300')  # 设置弹出框的大小 w x h

    l1 = tkinter.Label(z, text="请输入待办事项：")
    l1.pack()  # 这里的side可以赋值为LEFT  RTGHT TOP  BOTTOM
    var1 = tkinter.StringVar()  # 这即是输入框中的内容
    var1.set('')  # 通过var.get()/var.set() 来 获取/设置var的值
    entry1 = tkinter.Entry(z, textvariable=var1)  # 设置"文本变量"为var
    entry1.pack()  # 将entry"打上去"

    l2 = tkinter.Label(z, text="请输入截止时间（xx:xx形式）：")
    l2.pack()  # 这里的side可以赋值为LEFT  RIGHT TOP  BOTTOM
    var2 = tkinter.StringVar()  # 这即是输入框中的内容
    var2.set('')  # 通过var.get()/var.set() 来 获取/设置var的值
    entry2 = tkinter.Entry(z, textvariable=var2)  # 设置"文本变量"为var
    entry2.pack()  # 将entry"打上去"

    tkinter.Button(z, text='确定', command=lambda: input_int()).place(x=180, y=100)  # 按下此按钮(Input), 触发input_int函数
    tkinter.Button(z, text='清除', command=input_clear).place(x=230, y=100)  # 按下此按钮(Clear), 触发input_clear函数

    # 上述完成之后, 开始真正弹出弹出框
    z.mainloop()


def read_or_write(w_num, d_num):
    def cancel(x):
        x.destroy()

        def clear():
            workbook = openpyxl.load_workbook('待办事项.xlsx')
            sheet = workbook['待办事项']
            for i in checkboxes:  # 遍历checkboxes字典，元组类型为bool
                if checkboxes[i].get() == True:  # 如果当前元组为true，则清除excel中内容
                    sheet.cell(row=i + 1, column=(w_num - 1) * 7 + d_num).value = ' '

            workbook.save('待办事项.xlsx')
            print_class(w_num)

            # exit()

        # 开始构造界面
        global window
        window = tkinter.Toplevel()
        window.geometry('400x500+500+100')

        tol = 0
        for i in range(20):
            if info[i][(w_num - 1) * 7 + d_num - 1] != ' ':
                tol += 1

        # 创建字典，用于存储选择状态的bool对象
        checkboxes = {}

        # 使用for遍历运动项目字典，并动态创建Checkbutton
        for i in range(tol):
            checkboxes[i] = BooleanVar()  # 创建bool变量对象，添加到checkboxes字典

            # 创建Checkbutton复选按钮
            t = info[i][(w_num - 1) * 7 + d_num - 1]
            one = t.find('\n')
            t = t[:one] + " " + t[one + 1:]
            Checkbutton(window,
                        text=t,
                        variable=checkboxes[i]).grid(row=i + 1, sticky=W)  # 靠西分布，每创建一个另起一行

        Button(window,
               text="确定",
               width=10,
               command=clear).place(x=300, y=450)  # 按下按钮，触发打印事件

        window.mainloop()

    def read(x):
        x.destroy()

        t = ''
        for i in range(20):
            if info[i][(w_num - 1) * 7 + d_num - 1] != ' ':
                m = info[i][(w_num - 1) * 7 + d_num - 1]
                one = m.find('\n')
                t += m[:one] + "  " + m[one + 1:] + "\n"

        y = tkinter.Tk(className='邮学小帮手')  # 弹出框框名
        y.geometry('270x600+500+100')  # 设置弹出框的大小 w x h
        label = Label(y, text=t,
                      height=100, width=30, anchor='n', wraplength=200,
                      font=('微软雅黑', 10, 'bold'), justify='left',
                      padx=10, pady=10
                      )
        label.pack()  # 包装与定位组件，没有这句话的话标签不出现在x窗口
        y.mainloop()

    def write(x):
        x.destroy()
        pop_up_box(w_num, d_num)
        print_class(w_num)

    x = tkinter.Tk(className='邮学小帮手')  # 弹出框框名
    x.geometry('270x120+500+300')  # 设置弹出框的大小 w x h

    # 设置文本内容
    label = Label(x, text='选择功能:',
                  height=5, width=20, anchor='n', wraplength=200,
                  font=('微软雅黑', 10, 'bold'), justify='left',
                  padx=10, pady=20
                  )
    label.pack()  # 包装与定位组件，没有这句话的话标签不出现在x窗口

    tkinter.Button(x, text='读取', command=lambda: read(x)).place(x=40, y=70)  # 按下此按钮(Input), 触发read函数
    tkinter.Button(x, text='写入', command=lambda: write(x)).place(x=200, y=70)  # 按下此按钮(Clear), 触发write函数
    tkinter.Button(x, text='清除', command=lambda: cancel(x)).place(x=120, y=70)  # 按下此按钮(Input), 触发read函数

    # 上述完成之后, 开始真正弹出弹出框
    x.mainloop()


def read_info():
    workbook = openpyxl.load_workbook('待办事项.xlsx')
    sheet = workbook['待办事项']
    info_ = [[]]
    for r in range(20):
        for c in range(112):
            temp = copy.deepcopy(sheet.cell(row=r + 1, column=c + 1).value)
            info_[r].append(temp)
        info_.append([])
    return info_


def draw_window(pic_way):
    if pic_way:
        img_open = Image.open(pic_way)
        global img_png
        img_png = ImageTk.PhotoImage(img_open)
        cv.create_image(600, 350, image=img_png)
    # else:
    #     cv.create_image(600, 350)
    week_names = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
    class_times = ["1", "2", "3", "4", "5", "6"]

    for day in range(7):
        cv.create_text(150 + day * 160, 20, text=week_names[day], fill='#FF1493', font=('微软雅黑', 10, 'bold'))  # 星期文字
        for line in range(6):
            if day == 0:
                cv.create_text(50, 100 + line * 100, text=class_times[line], fill='#FF1493',
                               font=('微软雅黑', 10, 'bold'))  # 节次文字


def print_class(week):
    global t
    try:
        if not week:
            t_ = print_class(get_current_week())
            draw_buttons_2(get_current_week())
        else:
            for each_ in t:
                cv.delete(each_)
            t_ = [cv.create_text(50, 30, text="第{}周".format(week), fill='#FF1493', font=('微软雅黑', 15, 'bold'))]
            for day in range(7):
                for line in range(6):
                    if info[5][(week - 1) * 7 + day] == " ":
                        t_.append(
                            cv.create_text(150 + day * 160, 100 + line * 100, fill='#000000',
                                           text=info[line][(week - 1) * 7 + day],
                                           font=('微软雅黑', 10, 'bold')))  # 待办事项文字
                    elif info[5][(week - 1) * 7 + day] != " ":
                        if line < 5:
                            t_.append(
                                cv.create_text(150 + day * 160, 100 + line * 100, fill='#000000',
                                               text=info[line][(week - 1) * 7 + day],
                                               font=('微软雅黑', 10, 'bold')))  # 待办事项文字
                        if line == 5:
                            t_[5] = cv.create_text(150 + day * 160, 100 + line * 100, fill='#000000',
                                                   text='（更多待办点击详情观看）',
                                                   font=('微软雅黑', 10, 'bold'))  # 待办事项文字
            draw_buttons_2(week)
        return t_
    except:
        return 0


def draw_buttons_2(week):
    days_F = []
    for i in range(7):
        exec('''def day_{}():
                global d
                d = read_or_write({}, {})'''.format(i, week, i + 1))
        days_F.append("day_{}".format(i))
    temp = 0
    for each in days_F[:]:
        Button(root, text="查看明细", command=eval(each), fg='#FF7256',
               activeforeground='#CD6600',
               font=('微软雅黑', 10, 'bold')).place(x=120 + temp * 160, y=35)
        temp = temp + 1


def get_current_week():
    # 获取现在周数
    date0 = '2022-02-28'
    d0 = datetime.datetime.strptime(date0, '%Y-%m-%d')
    date1 = time.strftime("%Y-%m-%d", time.localtime())
    d1 = datetime.datetime.strptime(date1, '%Y-%m-%d')
    return int((d1 - d0).days / 7) + 1

def draw_buttons():
    weeks_F = []
    for i in range(17):
        exec('''def week_{}():
                global t
                t = print_class({})'''.format(i, i))
        weeks_F.append("week_{}".format(i))
    temp = 1
    Button(root, text="当前周".format(temp), command=eval('week_0'), fg='#FF7256', activeforeground='#CD6600',
           font=('微软雅黑', 18, 'bold')).place(x=1205, y=5)
    for each in weeks_F[1:]:
        Button(root, text="第{0:^5}周".format(temp), command=eval(each), fg='#FF7256', activeforeground='#CD6600',
               width=10, font=('微软雅黑', 10, 'bold')).place(x=1205, y=25 + temp * 40)
        temp = temp + 1

# def data_update():
#     info = read_info()
#     root.after(1000,data_update)

def show_calendar(pic_way):
    draw_window(pic_way)
    global t
    t = print_class(0)
    draw_buttons()
    cv.pack(anchor='w')
    # root.after(1000,data_update)
    root.mainloop()


def todo_main(pic_way=''):
    # 处理文件
    if not os.path.exists('./待办事项.xlsx'):
        workbook = openpyxl.Workbook()
        ws = workbook.create_sheet(title='待办事项')
        for i in range(20):
            for j in range(112):  #
                ws.cell(row=i + 1, column=j + 1).value = ' '
        workbook.save('待办事项.xlsx')  # 保存变更

    # create window
    global cv, root, t, img_png, info, t, d, window
    root = Toplevel()
    root.config(bg='#BFEFFF')
    root.title("邮学小帮手")
    root.geometry('1300x700+100+20')
    info = read_info()
    cv = Canvas(root, bg="#BFEFFF", width=1200, height=700)  # create canvas   #BFEFFF
    # 一些全局变量
    t = []
    d = []
    img_png = None
    show_calendar(pic_way)
    # show_calendar(pic_way='C:/Users/……')
